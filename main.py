import sys
import openpyxl as op
import math
import random
import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from matplotlib.ticker import MultipleLocator
from PySide6.QtWidgets import QApplication, QMainWindow
from matplotlib.backends.backend_qtagg import FigureCanvasQTAgg as FigureCanvas
from graf import Ui_MainWindow


class Fire(QMainWindow):
    def __init__(self):
        super(Fire, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)
        self.buttons()

    def buttons(self):
        self.ui.pushButton.clicked.connect(lambda: self.firedot0(int(self.ui.spinBox.text()),
                                                                 (int(self.ui.spinBox_2.text()))+4))
        self.ui.pushButton_2.clicked.connect(lambda: self.fireugl(int(self.ui.spinBox.text()),
                                                                  (int(self.ui.spinBox_2.text())) + 4))
        self.ui.pushButton_3.clicked.connect(lambda: self.firenar(int(self.ui.spinBox.text()),
                                                                  (int(self.ui.spinBox_2.text())) + 4))
        self.ui.pushButton_4.clicked.connect(lambda: self.firetl(int(self.ui.spinBox.text()),
                                                                 (int(self.ui.spinBox_2.text())) + 4))
        self.ui.pushButton_5.clicked.connect(lambda: self.firefasad())

        self.ui.pushButton.clicked.connect(lambda: self.func(int(self.ui.spinBox.text()), 'Режим ГОСТ 30247.0.xlsx')
                                           if str(self.ui.lineEdit_4.text()) == ''
                                           else self.func(int(self.ui.spinBox.text()),
                                                          str(self.ui.lineEdit_4.text()) + '.xlsx'))

        self.ui.pushButton_2.clicked.connect(lambda: self.func(int(self.ui.spinBox.text()), 'Углеводород.xlsx')
                                             if str(self.ui.lineEdit_4.text()) == ''
                                             else self.func(int(self.ui.spinBox.text()),
                                                            str(self.ui.lineEdit_4.text()) + '.xlsx'))

        self.ui.pushButton_3.clicked.connect(lambda: self.func(int(self.ui.spinBox.text()), 'Наружный режим.xlsx')
                                             if str(self.ui.lineEdit_4.text()) == ''
                                             else self.func(int(self.ui.spinBox.text()),
                                                            str(self.ui.lineEdit_4.text()) + '.xlsx'))

        self.ui.pushButton_4.clicked.connect(lambda: self.func(int(self.ui.spinBox.text()), 'Тлеющий режим.xlsx')
                                             if str(self.ui.lineEdit_4.text()) == ''
                                             else self.func(int(self.ui.spinBox.text()),
                                                            str(self.ui.lineEdit_4.text()) + '.xlsx'))
        self.ui.pushButton_5.clicked.connect(lambda: self.func(7, 'Фасад.xlsx'))

    def firedot0(self, a, b):
        book = op.Workbook()
        sheet = book.active
        sheet['B2'] = 'Время'
        sheet['C2'] = 'Тmin'
        sheet['D2'] = 'Tmax'
        sheet['E2'] = 'Tпож'
        for i in range(3, b):
            sheet.cell(row=i, column=2).value = i - 3
            sheet.cell(row=i, column=5).value = 345 * math.log10((i - 3) * 8 + 1) + 20
            if (i - 3) <= 10:
                sheet.cell(row=i, column=3).value = 0.85 * (345 * math.log10((i - 3) * 8 + 1) + 20)
                sheet.cell(row=i, column=4).value = 1.15 * (345 * math.log10((i - 3) * 8 + 1) + 20)
            elif 11 <= (i - 3) < 30:
                sheet.cell(row=i, column=3).value = 0.90 * (345 * math.log10((i - 3) * 8 + 1) + 20)
                sheet.cell(row=i, column=4).value = 1.10 * (345 * math.log10((i - 3) * 8 + 1) + 20)
            else:
                sheet.cell(row=i, column=3).value = 0.95 * (345 * math.log10((i - 3) * 8 + 1) + 20)
                sheet.cell(row=i, column=4).value = 1.05 * (345 * math.log10((i - 3) * 8 + 1) + 20)
        if a > 0:
            for u in range(a):
                sheet.cell(row=2, column=u+6).value = 'Т' + str(u+1)
                n = 0
                r = random.uniform(0.98, 1.02) \
                    * random.uniform(0.99, 1.01) \
                    * random.uniform(0.99, 1.01) \
                    * random.uniform(0.99, 1.01)
                for y in range(3, b):
                    if 1 < n < 5:
                        sheet.cell(row=y, column=u + 6).value = random.uniform(0.98, 1.02) * r \
                                                                * 345 * math.log10((y - 3) * 8 + 1) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) * 50\
                                                                + random.uniform(-10, 10)
                        n += 1
                    elif 5 <= n < 11:
                        sheet.cell(row=y, column=u + 6).value = r * 345 * math.log10((y - 3) * 8 + 1) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) \
                                                                * 15 + random.uniform(-7, 7)
                        n += 1
                    else:
                        n += 1
                        sheet.cell(row=y, column=u + 6).value = r * 345 * math.log10((y - 3) * 8 + 1) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) * 7

            sheet.cell(row=2, column=a+6).value = 'Тср'

            for t in range(3, b):
                o = 0
                for m in range(a):
                    o += sheet.cell(row=t, column=(a + 6) - (m + 1)).value
                sheet.cell(row=t, column=a + 6).value = o / a
        if str(self.ui.lineEdit_4.text()) == '':
            book.save('Режим ГОСТ 30247.0.xlsx')
        else:
            book.save(str(self.ui.lineEdit_4.text()) + '.xlsx')

    def fireugl(self, a, b):
        book = op.Workbook()
        sheet = book.active
        sheet['B2'] = 'Время'
        sheet['C2'] = 'Тmin'
        sheet['D2'] = 'Tmax'
        sheet['E2'] = 'Tпож'

        for i in range(3, b):
            sheet.cell(row=i, column=2).value = i - 3
            sheet.cell(row=i, column=5).value = 1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                        * math.e ** (-2.5 * (i - 3))) + 20
            if (i - 3) <= 10:
                sheet.cell(row=i, column=3).value = 0.85 * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                                    * math.e ** (-2.5 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = 1.15 * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                                    * math.e ** (-2.5 * (i - 3))) + 20)
            elif 10 < (i - 3) <= 30:
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (15 - 0.5 * ((i - 3) - 10)))) \
                                                    * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                               * math.e ** (-2.5 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (15 - 0.5 * ((i - 3) - 10)))) \
                                                    * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                               * math.e ** (-2.5 * (i - 3))) + 20)
            elif 30 < (i - 3) <= 60:
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (5 - 0.083 * ((i - 3) - 30)))) \
                                                    * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                               * math.e ** (-2.5 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (5 - 0.083 * ((i - 3) - 30)))) \
                                                    * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                               * math.e ** (-2.5 * (i - 3))) + 20)
            else:
                sheet.cell(row=i, column=3).value = 0.975 * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                                     * math.e ** (-2.5 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = 1.025 * (1080 * (1 - 0.325 * math.e ** (-0.167 * (i - 3)) - 0.675
                                                                     * math.e ** (-2.5 * (i - 3))) + 20)

        if a > 0:
            for u in range(a):
                sheet.cell(row=2, column=u + 6).value = 'Т' + str(u + 1)
                n = 0
                r = random.uniform(0.98, 1.02) * random.uniform(0.99, 1.01) * random.uniform(0.99, 1.01) \
                    * random.uniform(0.99, 1.01)
                for y in range(3, b):
                    if 1 < n < 5:
                        sheet.cell(row=y, column=u + 6).value = random.uniform(0.95, 1.01) * r * 1080 \
                                                                * (1 - 0.325 * math.e ** (-0.167 * (y - 3)) - 0.675
                                                                   * math.e ** (-2.5 * (y - 3))) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) \
                                                                * 40 + random.uniform(-10, 10)
                        n += 1
                    elif 5 <= n < 11:
                        sheet.cell(row=y, column=u + 6).value = random.uniform(0.95, 1.01) * r * 1080 \
                                                                * (1 - 0.325 * math.e ** (-0.167 * (y - 3)) - 0.675
                                                                   * math.e ** (-2.5 * (y - 3))) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) \
                                                                * 15 + random.uniform(-5, 5)
                        n += 1
                    else:
                        n += 1
                        sheet.cell(row=y, column=u + 6).value = r * 1080 * (1 - 0.325 * math.e ** (-0.167 * (y - 3))
                                                                            - 0.675 * math.e ** (-2.5 * (y - 3))) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) * 7
            sheet.cell(row=2, column=a + 6).value = 'Тср'

            for t in range(3, b):
                o = 0
                for m in range(a):
                    o += sheet.cell(row=t, column=(a + 6) - (m + 1)).value
                sheet.cell(row=t, column=a + 6).value = o / a
        if str(self.ui.lineEdit_4.text()) == '':
            book.save('Углеводород.xlsx')
        else:
            book.save(str(self.ui.lineEdit_4.text()) + '.xlsx')

    def firenar(self, a, b):
        book = op.Workbook()
        sheet = book.active
        sheet['B2'] = 'Время'
        sheet['C2'] = 'Тmin'
        sheet['D2'] = 'Tmax'
        sheet['E2'] = 'Tпож'

        for i in range(3, b):
            sheet.cell(row=i, column=2).value = i - 3
            sheet.cell(row=i, column=5).value = 660 * (1 - 0.687 * math.e **
                                                       (-0.32 * (i - 3)) - 0.313 * math.e ** (-3.8 * (i - 3))) + 20
            if (i - 3) <= 10:
                sheet.cell(row=i, column=3).value = 0.85 * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3))
                                                                   - 0.313 * math.e ** (-3.8 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = 1.15 * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3))
                                                                   - 0.313 * math.e ** (-3.8 * (i - 3))) + 20)
            elif 10 < (i - 3) <= 30:
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (15 - 0.5 * ((i - 3) - 10)))) \
                                                    * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3)) - 0.313
                                                              * math.e ** (-3.8 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (15 - 0.5 * ((i - 3) - 10)))) \
                                                    * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3)) - 0.313
                                                              * math.e ** (-3.8 * (i - 3))) + 20)
            elif 30 < (i - 3) <= 60:
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (5 - 0.083 * ((i - 3) - 30)))) \
                                                    * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3)) - 0.313
                                                              * math.e ** (-3.8 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (5 - 0.083 * ((i - 3) - 30)))) \
                                                    * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3)) - 0.313
                                                              * math.e ** (-3.8 * (i - 3))) + 20)
            else:
                sheet.cell(row=i, column=3).value = 0.975 * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3))
                                                                    - 0.313 * math.e ** (-3.8 * (i - 3))) + 20)
                sheet.cell(row=i, column=4).value = 1.025 * (660 * (1 - 0.687 * math.e ** (-0.32 * (i - 3))
                                                                    - 0.313 * math.e ** (-3.8 * (i - 3))) + 20)

        if a > 0:
            for u in range(a):
                sheet.cell(row=2, column=u + 6).value = 'Т' + str(u + 1)
                n = 0
                r = random.uniform(0.98, 1.02) * random.uniform(0.99, 1.01) * random.uniform(0.99, 1.01)
                for y in range(3, b):
                    if 1 < n < 5:
                        sheet.cell(row=y, column=u + 6).value = random.uniform(0.96, 1.01) * r * 660 * \
                                                                (1 - 0.687 * math.e ** (-0.32 * (y - 3)) - 0.313 *
                                                                 math.e ** (-3.8 * (y - 3))) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) \
                                                                * 30 + random.uniform(-7, 7)
                        n += 1
                    elif 5 <= n < 11:
                        sheet.cell(row=y, column=u + 6).value = random.uniform(0.96, 1.01) * r * 660 * \
                                                                (1 - 0.687 * math.e ** (-0.32 * (y - 3)) - 0.313 *
                                                                 math.e ** (-3.8 * (y - 3))) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) \
                                                                * 15 + random.uniform(-3, 3)
                        n += 1
                    else:
                        n += 1
                        sheet.cell(row=y, column=u + 6).value = r * 660 * (1 - 0.687 * math.e ** (-0.32 * (y - 3))
                                                                           - 0.313 * math.e ** (-3.8 * (y - 3))) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) * 7
            sheet.cell(row=2, column=a + 6).value = 'Тср'

            for t in range(3, b):
                o = 0
                for m in range(a):
                    o += sheet.cell(row=t, column=(a + 6) - (m + 1)).value
                sheet.cell(row=t, column=a + 6).value = o / a
        if str(self.ui.lineEdit_4.text()) == '':
            book.save('Наружный режим.xlsx')
        else:
            book.save(str(self.ui.lineEdit_4.text()) + '.xlsx')

    def firetl(self, a, b):
        book = op.Workbook()
        sheet = book.active
        sheet['B2'] = 'Время'
        sheet['C2'] = 'Тmin'
        sheet['D2'] = 'Tmax'
        sheet['E2'] = 'Tпож'

        for i in range(3, b):
            sheet.cell(row=i, column=2).value = i - 3
            if (i - 3) <= 10:
                sheet.cell(row=i, column=5).value = 154 * (i-3)**0.25 + 20
                sheet.cell(row=i, column=3).value = 0.85 * (154 * (i-3)**0.25 + 20)
                sheet.cell(row=i, column=4).value = 1.15 * (154 * (i-3)**0.25 + 20)
            elif 10 < (i - 3) <= 21:
                sheet.cell(row=i, column=5).value = 154 * (i-3)**0.25 + 20
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (15 - 0.5 * ((i - 3) - 10)))) * (154 * (i-3)**0.25
                                                                                                  + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (15 - 0.5 * ((i - 3) - 10)))) * (154 * (i-3)**0.25
                                                                                                  + 20)
            elif 21 < (i - 3) <= 30:
                sheet.cell(row=i, column=5).value = 345 * math.log10(8 * ((i - 3) - 20) + 1) + 20
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (15 - 0.5 * ((i - 3) - 10)))) \
                                                    * (345 * math.log10(8 * ((i - 3) - 20) + 1) + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (15 - 0.5 * ((i - 3) - 10)))) \
                                                    * (345 * math.log10(8 * ((i - 3) - 20) + 1) + 20)
            elif 30 < (i - 3) <= 60:
                sheet.cell(row=i, column=5).value = 345 * math.log10(8 * ((i - 3) - 20) + 1) + 20
                sheet.cell(row=i, column=3).value = (1 - (0.01 * (5 - 0.083 * ((i - 3) - 30)))) \
                                                    * (345 * math.log10(8 * ((i - 3) - 20) + 1) + 20)
                sheet.cell(row=i, column=4).value = (1 + (0.01 * (5 - 0.083 * ((i - 3) - 30)))) \
                                                    * (345 * math.log10(8 * ((i - 3) - 20) + 1) + 20)
            else:
                sheet.cell(row=i, column=5).value = 345 * math.log10(8 * ((i - 3) - 20) + 1) + 20
                sheet.cell(row=i, column=3).value = 0.975 * (345 * math.log10(8 * ((i - 3) - 20) + 1) + 20)
                sheet.cell(row=i, column=4).value = 1.025 * (345 * math.log10(8 * ((i - 3) - 20) + 1) + 20)

        if a > 0:
            for u in range(a):
                n = 0
                sheet.cell(row=2, column=u + 6).value = 'Т' + str(u + 1)
                r = random.uniform(0.98, 1.02) * random.uniform(0.99, 1.02) * random.uniform(0.99, 1.01) \
                    * random.uniform(0.99, 1.01)
                for y in range(3, b):
                    if 0 <= n < 5:
                        sheet.cell(row=y, column=u + 6).value = r * 154 * (y - 3)**0.25 + 20 + random.randint(-1, 1) \
                                                                * random.betavariate(1, 3) * 40 + random.uniform(-5, 5)
                        n += 1
                    elif 5 <= n <= 21:
                        sheet.cell(row=y, column=u + 6).value = r * 154 * (y - 3)**0.25 + 20 + random.randint(-1, 1) \
                                                                * random.betavariate(1, 3) * 10 + random.uniform(-3, 3)
                        n += 1
                    else:
                        sheet.cell(row=y, column=u + 6).value = r * 345 * math.log10(8 * ((y - 3) - 20) + 1) + 20 \
                                                                + random.randint(-1, 1) * random.betavariate(1, 3) * 5
                        n += 1
            sheet.cell(row=2, column=a + 6).value = 'Тср'

            for t in range(3, b):
                o = 0
                for m in range(a):
                    o += sheet.cell(row=t, column=(a + 6) - (m + 1)).value
                sheet.cell(row=t, column=a + 6).value = o / a
        if str(self.ui.lineEdit_4.text()) == '':
            book.save('Тлеющий режим.xlsx')
        else:
            book.save(str(self.ui.lineEdit_4.text()) + '.xlsx')

    def firefasad(self):
        book = op.Workbook()
        sheet = book.active
        sheet['B2'] = 'Время'
        sheet['C2'] = 'Тmin'
        sheet['D2'] = 'Tmax'
        sheet['E2'] = 'Tпож'
        r = random.uniform(0.98, 1.02) * random.uniform(0.99, 1.01) * random.uniform(0.99, 1.01) \
            * random.uniform(0.99, 1.01)
        r2 = random.uniform(0.93, 1.07) * random.uniform(0.99, 1.02) * random.uniform(0.99, 1.01) \
             * random.uniform(0.99, 1.01)
        r3 = random.uniform(0.93, 1.03) * random.uniform(0.99, 1.01) * random.uniform(0.99, 1.01) \
             * random.uniform(0.99, 1.01)
        ran2 = random.uniform(0.85, 0.95)
        ran3 = random.uniform(0.75, 0.85)
        ran4 = random.uniform(0.55, 0.65)
        ran5 = random.uniform(0.45, 0.55)
        ran6 = random.uniform(0.35, 0.45)
        ran7 = random.uniform(0.15, 0.25)
        for i in range(3, 49):
            sheet.cell(row=i, column=2).value = i - 3
            if (i - 3) <= 7:
                sheet.cell(row=i, column=12).value = ran7 * r2 * (105 * (i-3) + 115 + random.randint(-1, 1) *
                                                                  random.betavariate(1, 3) * 60 + random.uniform(-3, 3))
                sheet.cell(row=i, column=11).value = ran6 * r2 * (105 * (i-3) + 115 + random.randint(-1, 1) *
                                                                  random.betavariate(1, 3) * 60 + random.uniform(-3, 3))
                sheet.cell(row=i, column=10).value = ran5 * r2 * (105 * (i-3) + 115 + random.randint(-1, 1) *
                                                                  random.betavariate(1, 3) * 60 + random.uniform(-3, 3))
                sheet.cell(row=i, column=9).value = ran4 * r2 * (105 * (i-3) + 115 + random.randint(-1, 1) *
                                                                 random.betavariate(1, 3) * 60 + random.uniform(-3, 3))
                sheet.cell(row=i, column=8).value = ran3 * r2 * (105 * (i-3) + 115 + random.randint(-1, 1) *
                                                                 random.betavariate(1, 3) * 60 + random.uniform(-3, 3))
                sheet.cell(row=i, column=7).value = ran2 * r2 * (105 * (i-3) + 115 + random.randint(-1, 1) *
                                                                 random.betavariate(1, 3) * 60 + random.uniform(-3, 3))
                sheet.cell(row=i, column=6).value = r2 * (105 * (i-3) + 115) + random.randint(-1, 1) * \
                                                    random.betavariate(1, 3) * 60 + random.uniform(-3, 3)

                sheet.cell(row=i, column=5).value = 105 * (i-3) + 115
                sheet.cell(row=i, column=3).value = (105 * (i-3) + 115) * 0.8
                sheet.cell(row=i, column=4).value = (105 * (i-3) + 115) * 1.2
            elif 7 < (i - 3) <= 25:
                sheet.cell(row=i, column=12).value = ran7 * r * (850 + random.randint(-1, 1) * random.betavariate(1, 3)
                                                                 * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=11).value = ran6 * r * (850 + random.randint(-1, 1) * random.betavariate(1, 3)
                                                                 * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=10).value = ran5 * r * (850 + random.randint(-1, 1) * random.betavariate(1, 3)
                                                                 * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=9).value = ran4 * r * (850 + random.randint(-1, 1) * random.betavariate(1, 3)
                                                                * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=8).value = ran3 * r * (850 + random.randint(-1, 1) * random.betavariate(1, 3)
                                                                * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=7).value = ran2 * r * (850 + random.randint(-1, 1) * random.betavariate(1, 3)
                                                                * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=6).value = r * 850 + random.randint(-1, 1) * random.betavariate(1, 3) * 20 \
                                                    + random.uniform(-3, 3)

                sheet.cell(row=i, column=5).value = 850
                sheet.cell(row=i, column=3).value = 850 * 0.94
                sheet.cell(row=i, column=4).value = 850 * 1.06
            else:
                sheet.cell(row=i, column=12).value = ran7 * r3 * (850 - 20 * (i - 28) + random.randint(-1, 1) *
                                                                  random.betavariate(1, 3) * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=11).value = ran6 * r3 * (850 - 20 * (i - 28) + random.randint(-1, 1) *
                                                                  random.betavariate(1, 3) * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=10).value = ran5 * r3 * (850 - 20 * (i - 28) + random.randint(-1, 1) *
                                                                  random.betavariate(1, 3) * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=9).value = ran4 * r3 * (850 - 20 * (i - 28) + random.randint(-1, 1) *
                                                                 random.betavariate(1, 3) * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=8).value = ran3 * r3 * (850 - 20 * (i - 28) + random.randint(-1, 1) *
                                                                 random.betavariate(1, 3) * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=7).value = ran2 * r3 * (850 - 20 * (i - 28) + random.randint(-1, 1) *
                                                                 random.betavariate(1, 3) * 20 + random.uniform(-3, 3))
                sheet.cell(row=i, column=6).value = r3 * (850 - 20 * (i - 28)) + random.randint(-1, 1) * \
                                                    random.betavariate(1, 3) * 20 + random.uniform(-3, 3)

                sheet.cell(row=i, column=5).value = 850 - 20 * (i - 28)
                sheet.cell(row=i, column=3).value = (850 - 20 * (i - 28)) * 0.8
                sheet.cell(row=i, column=4).value = (850 - 20 * (i - 28)) * 1.2
        book.save('Фасад.xlsx')

    def func(self, a, b):
        plt.close('all')
        canvas = FigureCanvas(plt.figure(facecolor='#0c0c0c'))
        ax = plt.axes()
        ax.spines['bottom'].set_color('#dddddd')
        ax.spines['top'].set_color('#dddddd')
        ax.spines['right'].set_color('#dddddd')
        ax.spines['left'].set_color('#dddddd')
        ax.yaxis.label.set_color('#dddddd')
        ax.xaxis.label.set_color('#dddddd')
        ax.title.set_color('#dddddd')
        ax.yaxis.set_major_locator(MultipleLocator(base=100))
        ax.xaxis.set_major_locator(MultipleLocator(base=10))
        ax.tick_params(axis='x', colors='#dddddd')
        ax.tick_params(axis='y', colors='#dddddd')
        ax.set_facecolor('#0c0c0c')

        plt.minorticks_on()
        # включаем основную сетку
        plt.grid(which='major', linestyle=':')
        # включаем дополнительную сетку

        plt.tight_layout()

        self.ui.gridLayout_3.addWidget(canvas, 0, 0, 1, 1)

        df = pd.read_excel(b)
        data_array = df.to_numpy()
        data_array = data_array[1:]
        transposed_array = data_array.transpose()
        transposed_array = transposed_array[1:]
        x2, y2 = np.array(transposed_array[0]), np.array(transposed_array[1])
        x3, y3 = np.array(transposed_array[0]), np.array(transposed_array[2])
        x4, y4 = np.array(transposed_array[0]), np.array(transposed_array[3])
        plt.plot(x2, y2, color='blue')
        plt.plot(x3, y3, color='m')
        if a > 1:
            for i in range(a+1):
                plt.plot(np.array(transposed_array[0]), np.array(transposed_array[i+3]), color='white', linewidth=0.3)
                plt.plot(x4, y4, color='green', linestyle='dashed', linewidth=0.3)
        if b != 'Фасад.xlsx':
            x1, y1 = np.array(transposed_array[0]), np.array(transposed_array[a + 4])
            plt.plot(x1, y1, color='red')
        elif b == 'Фасад.xlsx':
            plt.plot([10, 20], [600, 600], color='white')
        plt.xlabel(' ')
        plt.ylabel(' ')


if __name__ == '__main__':
    app = QApplication(sys.argv)
    window = Fire()
    window.show()
    sys.exit(app.exec())
